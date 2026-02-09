# Jack Wilson
# 2/3/2026
# This file creates or updates the driver prediction list file used for prediction a specific race

# --------------------------------------------------------------------------------
# Import modules

import pandas as pd
import os, sys, time, warnings
import xlwings as xw
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

current_dir = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(current_dir))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

# --------------------------------------------------------------------------------
# Close Excel

def _close_workbook_xlwings(file_path):
    """
    If workbook is open in Excel, connect via xlwings, save, and close
    Only call when Excel is already running

    """
    # Try to close with xlwings
    path_abs = os.path.abspath(file_path)
    path_norm = os.path.normcase(path_abs)
    try:
        for app in xw.apps:
            for book in app.books:
                try:
                    fn = book.fullname
                    if fn and os.path.normcase(os.path.abspath(fn)) == path_norm:
                        book.save()
                        book.close()
                        time.sleep(0.3)
                        return True
                except Exception:
                    continue
    except Exception:
        pass
    return False


def _close_workbook_in_excel(file_path):
    """
    If workbook is open in Excel, close it so  file can be read/written
    Only runs when Excel is already running
    Tries xlwings first then win32com
    
    """
    xl = None
    try:
        xl = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return
    if xl is None:
        return
    if _close_workbook_xlwings(file_path):
        return
    target_abs = os.path.abspath(file_path)
    target_norm = os.path.normcase(target_abs)
    target_name = os.path.basename(file_path)
    try:
        target_real = os.path.realpath(target_abs)
    except OSError:
        target_real = target_abs
    try:
        if xl.Workbooks.Count == 0:
            return
    except Exception:
        return
    to_close = []
    try:
        count = xl.Workbooks.Count
    except Exception:
        return
    try:
        for i in range(1, count + 1):
            wb = xl.Workbooks(i)
            try:
                fn = wb.FullName
                if not fn:
                    if wb.Name == target_name:
                        to_close.append(wb.Name)
                        break
                    continue
                wb_abs = os.path.abspath(fn)
                wb_norm = os.path.normcase(wb_abs)
                try:
                    wb_real = os.path.realpath(wb_abs)
                except OSError:
                    wb_real = wb_abs
                if wb_norm == target_norm or wb_real == target_real:
                    to_close.append(wb.Name)
                    break
                if wb.Name == target_name:
                    to_close.append(wb.Name)
                    break
            except Exception:
                continue
        if not to_close:
            return
        xl.DisplayAlerts = False
        for name in to_close:
            try:
                xl.Workbooks(name).Close(SaveChanges=True)
            except Exception:
                pass
        time.sleep(0.3)
    except Exception:
        pass
    finally:
        try:
            if xl is not None:
                xl.DisplayAlerts = True
        except Exception:
            pass

# --------------------------------------------------------------------------------
# Create Excel

def create_driver_pred_list():
    """
    Create or overwrite driver_pred_list.xlsx with:
    - 'list' sheet: index 1-22, name/team/grand_prix columns with data validation from drivers/teams/races sheets
    - 'drivers' sheet: unique driver names from f1_data_pre_qual_clean.csv
    - 'teams' sheet: unique team names from f1_data_pre_qual_clean.csv
    - 'races' sheet: unique grand prix names from f1_data_pre_qual_clean.csv
    Driver, team, and race lists are refreshed from the CSV each time
    If the workbook already exists, existing name/team/grand_prix values in the "list" sheet are preserved
    
    """
    # Get file paths and laod data
    OUTPUT_PATH = os.path.join(PROJECT_ROOT, "driver_pred_list.xlsx")
    _close_workbook_in_excel(OUTPUT_PATH)
    time.sleep(0.5)
    file_existed = os.path.isfile(OUTPUT_PATH)
    csv_path = os.path.join(PROJECT_ROOT, "data", "final", "f1_data_pre_qual_clean.csv")
    df = pd.read_csv(csv_path, low_memory=False)

    # Get unique prediction values
    drivers = sorted(df["driver_name"].dropna().unique().tolist())
    teams   = sorted(df["team_name"].dropna().unique().tolist())
    races   = sorted(df["circuit_name"].dropna().unique().tolist())

    # Create workbook and all sheets
    wb = Workbook()
    wb.remove(wb.active)
    ws_list = wb.create_sheet("list", 0)
    ws_drivers = wb.create_sheet("drivers", 1)
    ws_teams   = wb.create_sheet("teams", 2)
    ws_races   = wb.create_sheet("races", 3)

    # Set up the 'list' sheet for user prediction input and appearance
    ws_list["A1"] = "index"
    ws_list["B1"] = "name"
    ws_list["C1"] = "team"
    ws_list["D1"] = "grand_prix"
    for i in range(1, 23):
        ws_list.cell(row=i + 1, column=1, value=i)
    default_width = 10
    for col in ["B", "C", "D"]:
        ws_list.column_dimensions[col].width = default_width * 2

    # Set up data validation for prediction columns
    n_drivers = len(drivers)
    n_teams = len(teams)
    n_races = len(races)
    dv_name = DataValidation(
        type="list",
        formula1=f"drivers!$A$1:$A${n_drivers}",
        allow_blank=True,
        showErrorMessage=False,
        showDropDown=False,
    )
    dv_team = DataValidation(
        type="list",
        formula1=f"teams!$A$1:$A${n_teams}",
        allow_blank=True,
        showErrorMessage=False,
        showDropDown=False,
    )
    dv_grand_prix = DataValidation(
        type="list",
        formula1=f"races!$A$1:$A${n_races}",
        allow_blank=True,
        showErrorMessage=False,
        showDropDown=False,
    )
    ws_list.add_data_validation(dv_name)
    ws_list.add_data_validation(dv_team)
    ws_list.add_data_validation(dv_grand_prix)
    for row in range(2, 24):
        dv_name.add(f"B{row}")
        dv_team.add(f"C{row}")
        dv_grand_prix.add(f"D{row}")
    default_width = 10
    ws_list.column_dimensions["B"].width = default_width * 2
    ws_list.column_dimensions["C"].width = default_width * 2
    ws_list.column_dimensions["D"].width = default_width * 2

    # Fill the 'drivers', 'teams', and 'races' sheets
    for r, name in enumerate(drivers, start=1):
        ws_drivers.cell(row=r, column=1, value=name)
    for r, name in enumerate(teams, start=1):
        ws_teams.cell(row=r, column=1, value=name)
    for r, name in enumerate(races, start=1):
        ws_races.cell(row=r, column=1, value=name)

    # Preserve existing predictions on 'list' if file already exists
    max_retries = 3
    retry_delay = 2
    for attempt in range(max_retries):
        try:
            if file_existed:
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", category=UserWarning, message=".*Data Validation.*")
                    existing_wb = load_workbook(OUTPUT_PATH, read_only=False, data_only=True)
                if "list" in existing_wb.sheetnames:
                    existing_list = existing_wb["list"]
                    for row in range(2, 24):
                        ws_list.cell(row=row, column=2, value=existing_list.cell(row=row, column=2).value)
                        ws_list.cell(row=row, column=3, value=existing_list.cell(row=row, column=3).value)
                        ws_list.cell(row=row, column=4, value=existing_list.cell(row=row, column=4).value)
                existing_wb.close()
            break
        except (PermissionError, OSError):
            if attempt < max_retries - 1:
                _close_workbook_in_excel(OUTPUT_PATH)
                time.sleep(retry_delay)
            continue
        except Exception:
            pass
            break

    # Save the workbook with fallback in case it cant overwrite
    for attempt in range(max_retries):
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, message=".*Data Validation.*")
                wb.save(OUTPUT_PATH)
            print("\nUpdated prediction list\n" if file_existed else "\nPrediction list created\n")
            return OUTPUT_PATH
        except (PermissionError, OSError):
            if attempt < max_retries - 1:
                _close_workbook_in_excel(OUTPUT_PATH)
                time.sleep(retry_delay)
            else:
                dir_name = os.path.dirname(OUTPUT_PATH)
                base_name = os.path.basename(OUTPUT_PATH)
                name_no_ext, ext = os.path.splitext(base_name)
                fallback_path = os.path.join(dir_name, f"{name_no_ext}_new{ext}")
                try:
                    with warnings.catch_warnings():
                        warnings.filterwarnings("ignore", category=UserWarning, message=".*Data Validation.*")
                        wb.save(fallback_path)
                    print(
                        "\nFile was open in Excel, so the updated list was saved to:\n"
                        f"  {fallback_path}\n"
                        "Close the original file in Excel, then replace it with this file\n"
                        "(or rename this file to driver_pred_list.xlsx).\n"
                    )
                    return fallback_path
                except Exception:
                    print(
                        "\nCould not save: file may be open in Excel or another program.\n"
                        "Please close driver_pred_list.xlsx and run the function again.\n"
                    )
                    raise


if __name__ == "__main__":
    create_driver_pred_list()
